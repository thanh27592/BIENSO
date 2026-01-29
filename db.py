import sqlite3
import threading
import os
from datetime import datetime


class DB:
    def __init__(self, path: str):
        self.path = path
        self.lock = threading.Lock()
        self.conn = sqlite3.connect(self.path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()

    def _init_schema(self):
        with self.lock:
            cur = self.conn.cursor()

            # Base tables (keep minimal columns for backward compatibility)
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    plate TEXT NOT NULL,
                    role TEXT NOT NULL,            -- IN / OUT
                    camera_name TEXT NOT NULL,
                    ts TEXT NOT NULL,              -- ISO string
                    confidence REAL,               -- ocr_conf (compat)
                    snapshot_path TEXT
                );
                """
            )

            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS sessions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    plate TEXT NOT NULL,
                    entry_time TEXT NOT NULL,
                    exit_time TEXT,                -- nullable
                    duration_sec INTEGER,          -- nullable
                    entry_event_id INTEGER,
                    exit_event_id INTEGER,
                    status TEXT NOT NULL           -- OPEN / CLOSED
                );
                """
            )

            # Indexes
            cur.execute("CREATE INDEX IF NOT EXISTS idx_events_plate_ts ON events(plate, ts);")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_plate_status ON sessions(plate, status);")

            # --- MIGRATION: add extra columns to events for tracking / debug ---
            self._ensure_column("events", "focus_id", "TEXT")
            self._ensure_column("events", "kind", "TEXT")
            self._ensure_column("events", "det_conf", "REAL")
            self._ensure_column("events", "ocr_conf", "REAL")
            self._ensure_column("events", "sharpness", "REAL")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_events_focus_id ON events(focus_id);")

            # --- MIGRATION: sessions.plate_key để ghép IN/OUT theo key nhưng vẫn hiển thị plate đẹp ---
            self._ensure_column("sessions", "plate_key", "TEXT")
            cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_platekey_status ON sessions(plate_key, status);")

            self.conn.commit()

    def _ensure_column(self, table: str, column: str, coldef: str):
        """
        Add column if missing. Safe for existing DB.
        coldef example: "TEXT", "INTEGER", "REAL"
        """
        cur = self.conn.cursor()
        cur.execute(f"PRAGMA table_info({table});")
        cols = [r[1] for r in cur.fetchall()]  # (cid,name,type,...)
        if column not in cols:
            cur.execute(f"ALTER TABLE {table} ADD COLUMN {column} {coldef};")
            self.conn.commit()

    def make_plate_key(self, plate: str) -> str:
        """
        Key để ghép IN/OUT kiểu kho: lấy 4 số cuối (giảm nhầm nhưng vẫn thực tế).
        Ví dụ 59V185558 -> "5558"
        """
        if not plate:
            return ""
        digits = "".join(ch for ch in str(plate) if ch.isdigit())
        return digits[-4:] if len(digits) >= 4 else ""

    def insert_event(
        self,
        plate: str,
        role: str,
        camera_name: str,
        ts: datetime,
        conf: float = 0.0,
        snapshot_path: str | None = None,
        *,
        focus_id: str | None = None,
        kind: str | None = None,
        det_conf: float | None = None,
        ocr_conf: float | None = None,
        sharpness: float | None = None,
    ) -> int:
        """
        Insert 1 row vào bảng events.

        Backward compatible:
        - `conf` (positional) vẫn được dùng nếu không truyền `ocr_conf`.
        """
        if ocr_conf is None:
            ocr_conf = float(conf or 0.0)
        if det_conf is None:
            det_conf = 0.0
        if sharpness is None:
            sharpness = 0.0

        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                "INSERT INTO events(plate, role, camera_name, ts, confidence, snapshot_path, focus_id, kind, det_conf, ocr_conf, sharpness)"
                " VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                (
                    plate,
                    role,
                    camera_name,
                    ts.isoformat(timespec="seconds"),
                    float(ocr_conf),
                    snapshot_path,
                    focus_id or "",
                    kind or "",
                    float(det_conf),
                    float(ocr_conf),
                    float(sharpness),
                ),
            )
            self.conn.commit()
            return int(cur.lastrowid)

    def update_event(
        self,
        event_id: int,
        *,
        plate: str | None = None,
        snapshot_path: str | None = None,
        kind: str | None = None,
        det_conf: float | None = None,
        ocr_conf: float | None = None,
        sharpness: float | None = None,
    ) -> None:
        """
        Update event row (dùng khi cùng 1 đối tượng nhưng OCR cải thiện -> update log thay vì insert log mới).
        Chỉ update các field != None.
        """
        sets = []
        vals = []

        if plate is not None:
            sets.append("plate=?")
            vals.append(plate)

        if snapshot_path is not None:
            sets.append("snapshot_path=?")
            vals.append(snapshot_path)

        if kind is not None:
            sets.append("kind=?")
            vals.append(kind)

        if det_conf is not None:
            sets.append("det_conf=?")
            vals.append(float(det_conf))

        if ocr_conf is not None:
            sets.append("confidence=?")  # giữ backward: confidence = ocr_conf
            vals.append(float(ocr_conf))
            sets.append("ocr_conf=?")
            vals.append(float(ocr_conf))

        if sharpness is not None:
            sets.append("sharpness=?")
            vals.append(float(sharpness))

        if not sets:
            return

        with self.lock:
            cur = self.conn.cursor()
            sql = "UPDATE events SET " + ", ".join(sets) + " WHERE id=?"
            vals.append(int(event_id))
            cur.execute(sql, tuple(vals))
            self.conn.commit()

    def update_session_plate(self, session_id: int, plate_final: str, plate_key: str) -> None:
        """
        Khi 1 đối tượng đã ghi log nhưng sau đó OCR ra biển tốt hơn, ta update session để UI hiển thị đúng.
        """
        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                "UPDATE sessions SET plate=?, plate_key=? WHERE id=?",
                (plate_final, plate_key, int(session_id)),
            )
            self.conn.commit()

    def process_warehouse_event(
        self,
        plate_final: str,
        plate_key: str,
        event_id: int,
        ts: datetime,
        min_out_delay_sec: int,
    ) -> tuple[str, int]:
        """
        Logic kiểu "kho": chỉ dựa trên plate_key + thời gian.
        - Nếu chưa có OPEN session cho plate_key => tạo OPEN (action=IN)
        - Nếu đã có OPEN session:
            + nếu đủ min_out_delay => CLOSE (action=OUT)
            + nếu chưa đủ => TOO_EARLY (action=TOO_EARLY) và chỉ update plate đẹp hơn

        Return: (action, session_id)
        """
        if not plate_key:
            plate_key = self.make_plate_key(plate_final)

        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                """
                SELECT id, entry_time
                FROM sessions
                WHERE plate_key=? AND status='OPEN'
                ORDER BY id DESC
                LIMIT 1
                """,
                (plate_key,),
            )
            row = cur.fetchone()

            if row is None:
                # create OPEN
                cur.execute(
                    """
                    INSERT INTO sessions(plate, plate_key, entry_time, status, entry_event_id)
                    VALUES(?,?,?,?,?)
                    """,
                    (plate_final, plate_key, ts.isoformat(timespec="seconds"), "OPEN", int(event_id)),
                )
                sid = int(cur.lastrowid)
                self.conn.commit()
                return "IN", sid

            # existing OPEN session
            sid = int(row["id"])
            try:
                entry_ts = datetime.fromisoformat(row["entry_time"])
            except Exception:
                entry_ts = ts

            delta = (ts - entry_ts).total_seconds()

            if delta >= float(min_out_delay_sec):
                # close
                cur.execute(
                    """
                    UPDATE sessions
                    SET exit_time=?, duration_sec=?, status='CLOSED',
                        plate=?, plate_key=?, exit_event_id=?
                    WHERE id=?
                    """,
                    (
                        ts.isoformat(timespec="seconds"),
                        int(delta),
                        plate_final,
                        plate_key,
                        int(event_id),
                        sid,
                    ),
                )
                self.conn.commit()
                return "OUT", sid

            # too early -> keep OPEN but update plate (tốt hơn)
            cur.execute(
                "UPDATE sessions SET plate=?, plate_key=? WHERE id=?",
                (plate_final, plate_key, sid),
            )
            self.conn.commit()
            return "TOO_EARLY", sid

    def get_recent_datalog(self, limit: int = 100):
        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                """
                SELECT
                    s.id,
                    s.plate,
                    s.entry_time AS time_in,
                    s.exit_time  AS time_out,
                    s.duration_sec,
                    s.status,
                    e1.snapshot_path AS snap_in,
                    e2.snapshot_path AS snap_out
                FROM sessions s
                LEFT JOIN events e1 ON e1.id = s.entry_event_id
                LEFT JOIN events e2 ON e2.id = s.exit_event_id
                ORDER BY s.id DESC
                LIMIT ?
                """,
                (int(limit),),
            )
            return [dict(r) for r in cur.fetchall()]

    def export_day_to_excel(self, day: str, out_path: str):
        """
        Export sessions của 1 ngày (YYYY-MM-DD) ra Excel.
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        if not day or len(day) != 10:
            raise ValueError("day must be YYYY-MM-DD")

        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                """
                SELECT
                    s.id,
                    s.plate,
                    s.entry_time AS time_in,
                    s.exit_time  AS time_out,
                    s.duration_sec,
                    s.status,
                    e1.snapshot_path AS snap_in,
                    e2.snapshot_path AS snap_out
                FROM sessions s
                LEFT JOIN events e1 ON e1.id = s.entry_event_id
                LEFT JOIN events e2 ON e2.id = s.exit_event_id
                WHERE substr(s.entry_time, 1, 10) = ?
                ORDER BY s.id ASC
                """,
                (day,),
            )
            rows = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Sessions {day}"
        ws.append(["ID", "Plate", "Time IN", "Time OUT", "Duration(sec)", "Status", "Snap IN", "Snap OUT"])

        for r in rows:
            ws.append(
                [
                    r["id"],
                    r["plate"],
                    r["time_in"],
                    r["time_out"],
                    r["duration_sec"],
                    r["status"],
                    r["snap_in"],
                    r["snap_out"],
                ]
            )

        # autosize
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18

        os_dir = os.path.dirname(out_path)
        if os_dir:
            os.makedirs(os_dir, exist_ok=True)
        wb.save(out_path)

    # backward alias
    def export_to_excel(self, excel_path: str):
        """
        Export toàn bộ sessions + events ra Excel (legacy).
        """
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        with self.lock:
            cur = self.conn.cursor()
            cur.execute(
                "SELECT id, plate, role, camera_name, ts, confidence, snapshot_path, focus_id, kind, det_conf, ocr_conf, sharpness FROM events ORDER BY id ASC"
            )
            events = cur.fetchall()

            cur.execute("SELECT id, plate, plate_key, entry_time, exit_time, duration_sec, status FROM sessions ORDER BY id ASC")
            sessions = cur.fetchall()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Events"
        ws1.append(["ID", "Plate", "Role", "Camera", "Time", "OCRConf", "Snapshot", "FocusID", "Kind", "DetConf", "Sharpness"])
        for e in events:
            ws1.append(
                [
                    e["id"],
                    e["plate"],
                    e["role"],
                    e["camera_name"],
                    e["ts"],
                    e["confidence"],
                    e["snapshot_path"],
                    e["focus_id"],
                    e["kind"],
                    e["det_conf"],
                    e["sharpness"],
                ]
            )

        ws2 = wb.create_sheet("Sessions")
        ws2.append(["ID", "Plate", "PlateKey", "Time IN", "Time OUT", "Duration(sec)", "Status"])
        for s in sessions:
            ws2.append([s["id"], s["plate"], s["plate_key"], s["entry_time"], s["exit_time"], s["duration_sec"], s["status"]])

        # autosize
        for ws in (ws1, ws2):
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18

        os_dir = os.path.dirname(excel_path)
        if os_dir:
            os.makedirs(os_dir, exist_ok=True)
        wb.save(excel_path)
